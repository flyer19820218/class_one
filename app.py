import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import io
import re
from matplotlib import font_manager
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

st.set_page_config(page_title="801 專屬成績大數據主控台", layout="centered", page_icon="🎓")

# ==========================================
# 區塊 1: 登入系統狀態初始化
# ==========================================
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'role' not in st.session_state:
    st.session_state.role = None
if 'user_id' not in st.session_state:
    st.session_state.user_id = None
if 'passwords' not in st.session_state:
    st.session_state.passwords = {}

# ==========================================
# 區塊 2: 本地字體初始化
# ==========================================
FONT_NAME = "NotoSansTC-Regular.ttf"

if os.path.exists(FONT_NAME):
    font_manager.fontManager.addfont(FONT_NAME)
    dynamic_font_name = font_manager.FontProperties(fname=FONT_NAME).get_name()
    plt.rcParams['font.sans-serif'] = [dynamic_font_name, 'Microsoft JhengHei', 'Arial']
    plt.rcParams['axes.unicode_minus'] = False
else:
    st.warning("⚠️ 找不到字體檔！圖表的中文可能會變成方塊。請確認 NotoSansTC-Regular.ttf 與 app.py 放在同一個資料夾。")

def get_google_sheet_csv_url(url):
    try:
        if "export?format=csv" in url: return url
        match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
        if match:
            gid_match = re.search(r'gid=([0-9]+)', url)
            gid_param = f"&gid={gid_match.group(1)}" if gid_match else ""
            return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv{gid_param}"
    except: pass
    return None

@st.cache_data(ttl=60)
def load_data(url):
    csv_url = get_google_sheet_csv_url(url)
    if not csv_url: return None
    try:
        df = pd.read_csv(csv_url)
        df.columns = df.columns.astype(str).str.strip()
        return df
    except: return None

def check_is_grade_2_or_3(exam_name):
    if any(keyword in str(exam_name) for keyword in ["二上", "二下", "三上", "三下"]):
        return True
    return False

# 🌟 核心防彈機制：無腦鎖定第一欄，智慧辨識學生代碼
def load_and_standardize_csv(url):
    df = load_data(url)
    if df is None or df.empty: return None
    df = df.copy()
    
    if '英語' in df.columns: df = df.rename(columns={'英語': '英文'})
    
    # 絕對防呆：直接抓取第一欄作為學號判斷基準
    first_col = df.columns[0]
    df['__id_num'] = pd.to_numeric(df[first_col], errors='coerce')
    
    # 僅篩選出真正的 5 位數學生代號 (7xxxx ~ 9xxxx)
    df = df[df['__id_num'].notna() & (df['__id_num'] >= 70000) & (df['__id_num'] <= 99999)].copy()
    df['座號'] = df['__id_num'].astype(int)
    
    return df

# ==========================================
# 區塊 3: 導師功能 (全校總平均)
# ==========================================
def generate_semester_average_excel(df_menu):
    all_data = []
    
    for _, row in df_menu.iterrows():
        exam_name = str(row.get('考試檔案名稱或別名', row.get('考試名稱', ''))).strip()
        file_url = str(row.get('該次成績單的 Google 試算表網址', '')).strip()
        if not file_url: continue
        
        df_exam = load_and_standardize_csv(file_url)
        if df_exam is not None:
            is_grade_2_3 = check_is_grade_2_or_3(exam_name)
            
            for col in ['國文', '英文', '數學', '自然', '社會', '歷史', '地理', '公民']:
                if col not in df_exam.columns: df_exam[col] = np.nan
                else: df_exam[col] = pd.to_numeric(df_exam[col], errors='coerce')
            
            if is_grade_2_3:
                df_exam['社會_融合'] = df_exam[['歷史', '地理', '公民']].mean(axis=1).fillna(df_exam['社會'])
            else:
                df_exam['社會_融合'] = df_exam['社會']
            
            for _, s_row in df_exam.iterrows():
                record = {
                    '座號': int(s_row['座號']), # 5 位數班級代碼 (如 70101)
                    '姓名': str(s_row.get('姓名', '')).strip(),
                    '國文': s_row.get('國文', np.nan),
                    '英文': s_row.get('英文', np.nan),
                    '數學': s_row.get('數學', np.nan),
                    '自然': s_row.get('自然', np.nan),
                    '社會_融合': s_row.get('社會_融合', np.nan),
                    '歷史': s_row.get('歷史', np.nan) if is_grade_2_3 else s_row.get('社會', np.nan),
                    '地理': s_row.get('地理', np.nan) if is_grade_2_3 else s_row.get('社會', np.nan),
                    '公民': s_row.get('公民', np.nan) if is_grade_2_3 else s_row.get('社會', np.nan)
                }
                all_data.append(record)
                
    if not all_data: return None
        
    df_all = pd.DataFrame(all_data)
    df_mean = df_all.groupby(['座號', '姓名']).mean(numeric_only=True).reset_index()
    
    for col in ['國文', '英文', '數學', '自然', '社會_融合', '歷史', '地理', '公民']:
        if col not in df_mean.columns: df_mean[col] = np.nan
            
    df_mean['五科總平均'] = (df_mean['國文'].fillna(0) + df_mean['英文'].fillna(0) + 
                          df_mean['數學'].fillna(0) + df_mean['自然'].fillna(0) + 
                          df_mean['社會_融合'].fillna(0)) / 5
                          
    mask_all_na = df_mean[['國文', '英文', '數學', '自然', '社會_融合']].isna().all(axis=1)
    df_mean.loc[mask_all_na, '五科總平均'] = np.nan
    df_mean = df_mean.sort_values('座號')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "全校學期總平均"
    
    headers = ["班級座號(學號)", "姓名", "國文平均", "英文平均", "數學平均", "自然平均", "社會(融合)平均", "歷史平均", "地理平均", "公民平均", "五科總和平均"]
    ws.append(headers)
    
    for _, r in df_mean.iterrows():
        row_data = [
            r.get('座號', ''), r.get('姓名', ''),
            round(r.get('國文', np.nan), 2) if pd.notna(r.get('國文')) else "", 
            round(r.get('英文', np.nan), 2) if pd.notna(r.get('英文')) else "",
            round(r.get('數學', np.nan), 2) if pd.notna(r.get('數學')) else "", 
            round(r.get('自然', np.nan), 2) if pd.notna(r.get('自然')) else "",
            round(r.get('社會_融合', np.nan), 2) if pd.notna(r.get('社會_融合')) else "", 
            round(r.get('歷史', np.nan), 2) if pd.notna(r.get('歷史')) else "",
            round(r.get('地理', np.nan), 2) if pd.notna(r.get('地理')) else "", 
            round(r.get('公民', np.nan), 2) if pd.notna(r.get('公民')) else "",
            round(r.get('五科總平均', np.nan), 2) if pd.notna(r.get('五科總平均')) else ""
        ]
        ws.append(row_data)
        
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1: cell.font = Font(bold=True)
                
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 14
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def pr_to_level(pr):
    if pd.isna(pr): return ""
    if pr >= 91: return "A++"
    elif pr >= 85: return "A+"
    elif pr >= 75: return "A"
    elif pr >= 60: return "B++"
    elif pr >= 45: return "B+"
    elif pr >= 15: return "B"
    else: return "C"

# 🌟 精準會考落點結算引擎 (字串強制切割對齊版)
def generate_pr_report_excel(df_menu, target_class_code):
    all_data = []
    target_class_code = str(target_class_code).strip()
    
    for _, row in df_menu.iterrows():
        exam_name = str(row.get('考試檔案名稱或別名', row.get('考試名稱', ''))).strip()
        file_url = str(row.get('該次成績單的 Google 試算表網址', '')).strip()
        if not file_url: continue
        
        df_exam = load_and_standardize_csv(file_url)
        if df_exam is not None:
            is_grade_2_3 = check_is_grade_2_or_3(exam_name)
            
            for col in ['國文', '英文', '數學', '自然', '社會', '歷史', '地理', '公民']:
                if col not in df_exam.columns: df_exam[col] = np.nan
                else: df_exam[col] = pd.to_numeric(df_exam[col], errors='coerce')
                
            if is_grade_2_3:
                df_exam['社會_融合'] = df_exam[['歷史', '地理', '公民']].mean(axis=1).fillna(df_exam['社會'])
            else:
                df_exam['社會_融合'] = df_exam['社會']
                
            # 依據全校成績算 PR
            for sub in ['國文', '英文', '數學', '自然', '社會_融合']:
                df_exam[f'{sub}_PR'] = df_exam[sub].rank(pct=True) * 100
                
            for _, s_row in df_exam.iterrows():
                full_id_str = str(int(s_row['座號'])) # 5 位數 (如 '70101')
                if len(full_id_str) >= 4:
                    class_id = full_id_str[:-2] # 取前幾碼當作班級 (如 '701')
                    seat_id = int(full_id_str[-2:]) # 取最後兩碼當作座號 (如 1)
                    
                    # 篩選邏輯：如果輸入 01，則 701, 801 皆吻合；若輸入 701 則精準吻合
                    if class_id.endswith(target_class_code) or class_id == target_class_code:
                        all_data.append({
                            '學號': full_id_str,
                            '座號': seat_id,
                            '姓名': str(s_row.get('姓名', '')).strip(),
                            '國文PR': s_row.get('國文_PR', np.nan),
                            '英文PR': s_row.get('英文_PR', np.nan),
                            '數學PR': s_row.get('數學_PR', np.nan),
                            '自然PR': s_row.get('自然_PR', np.nan),
                            '社會PR': s_row.get('社會_融合_PR', np.nan)
                        })

    if not all_data: return None, None
    
    df_all = pd.DataFrame(all_data)
    df_mean = df_all.groupby(['學號', '座號', '姓名']).mean(numeric_only=True).reset_index()
    df_mean['五科總PR平均'] = df_mean[['國文PR', '英文PR', '數學PR', '自然PR', '社會PR']].mean(axis=1)
    
    for col in ['國文PR', '英文PR', '數學PR', '自然PR', '社會PR', '五科總PR平均']:
        df_mean[f'{col.replace("PR", "").replace("平均", "")}預估等級'] = df_mean[col].apply(pr_to_level)
        
    df_mean = df_mean.sort_values('學號')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"會考預估落點"
    
    headers = ["班級座號(學號)", "姓名", "國文PR", "國文預估", "英文PR", "英文預估", "數學PR", "數學預估", "自然PR", "自然預估", "社會PR", "社會預估", "五科平均PR", "綜合預估等級"]
    ws.append(headers)
    
    for _, r in df_mean.iterrows():
        ws.append([
            r.get('學號', ''), r.get('姓名', ''),
            round(r.get('國文PR', np.nan), 2) if pd.notna(r.get('國文PR')) else "", r.get('國文預估等級', ''),
            round(r.get('英文PR', np.nan), 2) if pd.notna(r.get('英文PR')) else "", r.get('英文預估等級', ''),
            round(r.get('數學PR', np.nan), 2) if pd.notna(r.get('數學PR')) else "", r.get('數學預估等級', ''),
            round(r.get('自然PR', np.nan), 2) if pd.notna(r.get('自然PR')) else "", r.get('自然預估等級', ''),
            round(r.get('社會PR', np.nan), 2) if pd.notna(r.get('社會PR')) else "", r.get('社會預估等級', ''),
            round(r.get('五科總PR平均', np.nan), 2) if pd.notna(r.get('五科總PR平均')) else "", r.get('五科總預估等級', '')
        ])
        
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1: cell.font = Font(bold=True)
            
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 13
    
    out = io.BytesIO()
    wb.save(out)
    return df_mean, out.getvalue()

# ==========================================
# 區塊 4: 雷達圖產生器 & 歷史數據撈取
# ==========================================
def draw_web_radar_chart(labels, pr_scores):
    pr_scores = [0 if pd.isna(x) else x for x in pr_scores]
    num_vars = len(labels)
    if num_vars == 0: return plt.figure()
        
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    plot_scores = list(pr_scores) + [pr_scores[0]]
    angles = angles + [angles[0]]
    
    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    
    display_labels = [f"{label}\n(PR)" for label in labels]
    ax.set_thetagrids(np.degrees(angles[:-1]), display_labels, fontsize=10)
    
    ax.set_ylim(0, 100)
    ax.set_yticks([25, 50, 75, 100])
    ax.set_yticklabels(["25", "50", "75", "100"], color="grey", size=8)
    
    ax.plot(angles, [50]*len(angles), color='#C0504D', linewidth=1.5, linestyle='--', label='PR 50 (中位數)')
    ax.plot(angles, plot_scores, color='#4F81BD', linewidth=2, linestyle='solid', label='個人表現')
    ax.fill(angles, plot_scores, color='#4F81BD', alpha=0.35)
    plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), fontsize=9, ncol=2)
    
    return fig

def fetch_student_history(df_menu, seat_num):
    history_records = []
    
    for _, row in df_menu.iterrows():
        exam_name = str(row.get('考試檔案名稱或別名', row.get('考試名稱', ''))).strip()
        file_url = str(row.get('該次成績單的 Google 試算表網址', '')).strip()
        if not file_url: continue
        
        df_exam = load_and_standardize_csv(file_url)
        if df_exam is not None:
            is_grade_2_3 = check_is_grade_2_or_3(exam_name)
            
            for col in ['國文', '英文', '數學', '自然', '社會', '歷史', '地理', '公民']:
                if col not in df_exam.columns: df_exam[col] = np.nan
                else: df_exam[col] = pd.to_numeric(df_exam[col], errors='coerce')
            
            if is_grade_2_3:
                df_exam['社會_融合'] = df_exam[['歷史', '地理', '公民']].mean(axis=1).fillna(df_exam['社會'])
            else:
                df_exam['社會_融合'] = df_exam['社會']
            
            for sub in ['國文', '英文', '數學', '自然', '社會_融合']:
                df_exam[f'{sub}_PR'] = df_exam[sub].rank(pct=True) * 100
            
            df_exam['總分'] = df_exam[['國文', '英文', '數學', '自然', '社會_融合']].sum(axis=1, skipna=True)
            df_exam['名次'] = df_exam['總分'].rank(ascending=False, method='min')
            
            match = df_exam[df_exam['座號'] == seat_num]
            if not match.empty:
                s_row = match.iloc[0]
                history_records.append({
                    "考試名稱": exam_name,
                    "國文": s_row.get('國文_PR', np.nan),
                    "英文": s_row.get('英文_PR', np.nan),
                    "數學": s_row.get('數學_PR', np.nan),
                    "自然": s_row.get('自然_PR', np.nan),
                    "社會(含領域融合)": s_row.get('社會_融合_PR', np.nan),
                    "總分": pd.to_numeric(s_row.get('總分', 0)),
                    "名次": pd.to_numeric(s_row.get('名次', 0))
                })
                
    return pd.DataFrame(history_records)

# ==========================================
# 區塊 5: 登入介面邏輯
# ==========================================
if not st.session_state.logged_in:
    st.title("🔐 801 專屬成績系統 - 登入")
    st.markdown("---")
    
    login_role = st.radio("請選擇您的身分：", ["導師", "學生", "家長"])
    
    if login_role == "導師":
        username = st.text_input("導師帳號：")
    else:
        username = st.text_input("請輸入班級座號 (5位數，例如 70101)：")
        
    password = st.text_input("請輸入密碼 (學生/家長預設密碼為座號)：", type="password")
    
    if st.button("🚀 登入系統", type="primary"):
        if login_role == "導師":
            if username == "yen" and password == "Axlc1982":
                st.session_state.logged_in = True
                st.session_state.role = "teacher"
                st.session_state.user_id = "yen"
                st.rerun()
            else:
                st.error("❌ 導師帳號或密碼錯誤！")
        else:
            try:
                seat_num = int(username)
                user_key = f"{login_role}_{seat_num}"
                expected_password = st.session_state.passwords.get(user_key, str(seat_num))
                
                if password == expected_password:
                    st.session_state.logged_in = True
                    st.session_state.role = "student" if login_role == "學生" else "parent"
                    st.session_state.user_id = seat_num
                    st.rerun()
                else:
                    st.error("❌ 密碼錯誤！")
            except ValueError:
                st.error("❌ 學生或家長帳號請輸入 5 位數純數字！")
    st.stop()

# ==========================================
# 區塊 6: 系統主介面 (登入後)
# ==========================================
col_title, col_logout = st.columns([0.8, 0.2])
with col_title:
    st.title("🎓 801 專屬成績大數據主控台")
with col_logout:
    role_name = "導師" if st.session_state.role == "teacher" else f"{'學生' if st.session_state.role == 'student' else '家長'} ({st.session_state.user_id})"
    st.info(f"👤 {role_name}")
    if st.button("登出"):
        st.session_state.logged_in = False
        st.session_state.role = None
        st.session_state.user_id = None
        st.rerun()

if st.session_state.role != "teacher":
    with st.expander("🔑 修改個人密碼"):
        new_pass = st.text_input("請輸入新密碼：", type="password")
        confirm_pass = st.text_input("再次確認新密碼：", type="password")
        if st.button("確認修改"):
            if new_pass == confirm_pass and new_pass != "":
                user_key = f"{'學生' if st.session_state.role == 'student' else '家長'}_{st.session_state.user_id}"
                st.session_state.passwords[user_key] = new_pass
                st.success("✅ 密碼修改成功！下次請使用新密碼登入。")
            else:
                st.error("❌ 兩次輸入的密碼不一致，或密碼不能為空。")

st.markdown("---")

MASTER_MENU_URL = "https://docs.google.com/spreadsheets/d/1FL-orK8H_oDrLuDg1pAijJICjHzPxJLcPIEngF0wko8/edit?gid=0#gid=0"
df_menu = load_data(MASTER_MENU_URL)

if df_menu is not None:
    if st.session_state.role == "teacher":
        with st.expander("📊 導師專屬：匯出所有學生學期各科總平均 Excel", expanded=False):
            st.write("系統將自動讀取主控台內的所有段考成績，計算全校每位學生的「各科跨學期平均」與「五科總平均」。")
            if st.button("🚀 結算並匯出全校總表", type="primary"):
                with st.spinner("正在結算各次段考數據中，請稍候..."):
                    excel_data = generate_semester_average_excel(df_menu)
                    if excel_data:
                        st.success("✅ 學期總平均計算完成！")
                        st.download_button(
                            label="📥 下載【全校學期總平均 Excel】",
                            data=excel_data,
                            file_name="全校學期總平均總表.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error("❌ 運算失敗，請確認主控台內是否有成績資料。")
        st.markdown("---")

    if '考試檔案名稱或別名' in df_menu.columns:
        exam_options = df_menu['考試檔案名稱或別名'].dropna().tolist()
    else:
        exam_options = df_menu['考試名稱'].dropna().tolist()
    
    tab1, tab2, tab3 = st.tabs(["🔍 單次段考詳細成績", "📈 歷史軌跡動態分析", "🎯 會考落點預估 (PR分析)"])
    
    with tab1:
        selected_exam = st.selectbox("📅 請選擇要查看的考試項目：", exam_options, key="student_exam")
        if '考試檔案名稱或別名' in df_menu.columns:
            target_sheet_url = df_menu[df_menu['考試檔案名稱或別名'] == selected_exam]['該次成績單的 Google 試算表網址'].values[0]
        else:
            target_sheet_url = df_menu[df_menu['考試名稱'] == selected_exam]['該次成績單的 Google 試算表網址'].values[0]
            
        has_7_subjects = check_is_grade_2_or_3(selected_exam)
        raw_df = load_data(target_sheet_url)
        
        if raw_df is not None:
            raw_df.columns = raw_df.columns.astype(str).str.strip()
            if '英語' in raw_df.columns: raw_df = raw_df.rename(columns={'英語': '英文'})
                
            id_col = None
            for col in raw_df.columns:
                vals = pd.to_numeric(raw_df[col], errors='coerce')
                if vals.notna().any() and ((vals >= 70000) & (vals <= 99999)).any():
                    id_col = col; break
            if id_col is None: id_col = raw_df.columns[0]
                
            raw_df['__id_num'] = pd.to_numeric(raw_df[id_col], errors='coerce')
            
            df_students = raw_df[raw_df['__id_num'].notna() & (raw_df['__id_num'] >= 70000) & (raw_df['__id_num'] <= 99999)].copy()
            df_students['座號'] = df_students['__id_num'].astype(int)
            df_bottom = raw_df[raw_df['__id_num'].isna() | (raw_df['__id_num'] < 70000) | (raw_df['__id_num'] > 99999)].copy()
            
            stats_dict = {}
            cols_to_extract = ['國文', '英文', '數學', '自然', '社會', '歷史', '地理', '公民']
            for _, row in df_bottom.iterrows():
                name = str(row.get('座號', row.get('班級座號', ''))).strip() if pd.notna(row.get('座號')) else str(row.get('姓名', '')).strip()
                if '高標' in name:
                    for c in cols_to_extract:
                        if c in row:
                            if c not in stats_dict: stats_dict[c] = {}
                            stats_dict[c]['高標'] = pd.to_numeric(row[c], errors='coerce')
                if '平均' in name or '均標' in name:
                    for c in cols_to_extract:
                        if c in row:
                            if c not in stats_dict: stats_dict[c] = {}
                            stats_dict[c]['平均'] = pd.to_numeric(row[c], errors='coerce')
            
            if has_7_subjects:
                display_subs = ['國文', '英文', '數學', '自然', '社會', '歷史', '地理', '公民']
                radar_subs = ['國文', '英文', '數學', '自然', '歷史', '地理', '公民']
            else:
                display_subs = ['國文', '英文', '數學', '自然', '社會']
                radar_subs = ['國文', '英文', '數學', '自然', '社會']
            
            for col in cols_to_extract:
                if col not in df_students.columns: df_students[col] = np.nan
            
            if has_7_subjects:
                df_students['總分'] = (pd.to_numeric(df_students['國文'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['英文'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['數學'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['自然'], errors='coerce').fillna(0) + 
                                      (pd.to_numeric(df_students['歷史'], errors='coerce').fillna(0) + 
                                       pd.to_numeric(df_students['地理'], errors='coerce').fillna(0) + 
                                       pd.to_numeric(df_students['公民'], errors='coerce').fillna(0)) / 3)
            else:
                df_students['總分'] = (pd.to_numeric(df_students['國文'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['英文'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['數學'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['自然'], errors='coerce').fillna(0) + 
                                      pd.to_numeric(df_students['社會'], errors='coerce').fillna(0))
            
            df_students['總PR'] = df_students['總分'].rank(pct=True) * 100
            
            if '名次' in df_students.columns:
                df_students['名次'] = pd.to_numeric(df_students['名次'], errors='coerce').fillna(0)
            else:
                df_students['名次'] = df_students['總分'].rank(ascending=False, method='min')
            
            for sub in radar_subs:
                df_students[sub] = pd.to_numeric(df_students[sub], errors='coerce').fillna(0)
                df_students[f'{sub}_PR_current'] = df_students[sub].rank(pct=True) * 100
            
            df_students = df_students.sort_values(by='座號').reset_index(drop=True)
            student_seats = sorted(df_students[df_students['座號'] > 0]['座號'].tolist())
            
            if st.session_state.role == "teacher":
                selected_seat = st.selectbox("🔢 請選擇要查詢的座號 (5位數，例如 70101)：", student_seats, key="student_seat")
            else:
                selected_seat = st.session_state.user_id
            
            if selected_seat in student_seats:
                stud_data = df_students[df_students['座號'] == selected_seat].iloc[0]
                st.markdown(f"### 👤 座號 {selected_seat} 號 【{stud_data.get('姓名', '')}】 的成績報告")
                
                col1, col2 = st.columns([1.1, 1.0])
                with col1:
                    st.write("**📊 各科成績明細：**")
                    for s in display_subs:
                        if s in stud_data and pd.notna(stud_data[s]):
                            score = float(stud_data[s])
                            stat = stats_dict.get(s, {})
                            avg = float(stat.get('平均', 0)) if pd.notna(stat.get('平均')) else 0.0
                            high = float(stat.get('高標', 0)) if pd.notna(stat.get('高標')) else 0.0
                            
                            if s in ['歷史', '地理', '公民']:
                                st.text(f"  └ {s}：{score:.2f}  (班均: {avg:.2f} | 高標: {high:.2f})")
                            else:
                                st.text(f"  {s}：{score:.2f}  (班均: {avg:.2f} | 高標: {high:.2f})")
                    
                    st.markdown("---")
                    total_score = float(stud_data['總分'])
                    rank = int(float(stud_data['名次']))
                    total_pr = float(stud_data['總PR'])
                    
                    st.metric(label="五科總分", value=f"{total_score:.2f}")
                    st.write(f"🏆 **排名**：第 {rank} 名")
                    st.write(f"📈 **總分 PR 值**：{total_pr:.2f}")
                    
                with col2:
                    pr_scores = [stud_data[f'{s}_PR_current'] for s in radar_subs]
                    fig = draw_web_radar_chart(radar_subs, pr_scores)
                    st.pyplot(fig)
            else:
                st.warning("⚠️ 此次考試中查無此座號成績資料。")
        else:
            st.warning("無法載入此考試資料。")

    with tab2:
        st.markdown("### 📊 學生歷史跨學期趨勢追蹤")
        if 'student_seats' in locals() and student_seats:
            if st.session_state.role == "teacher":
                analysis_seat = st.selectbox("🔢 請選擇要分析的學生座號 (5位數)：", student_seats, key="analysis_seat")
            else:
                analysis_seat = st.session_state.user_id
            
            if analysis_seat in student_seats:
                with st.spinner("正在為您統整歷史數據..."):
                    df_history = fetch_student_history(df_menu, analysis_seat)
                    
                if not df_history.empty:
                    stud_name = df_students[df_students['座號'] == analysis_seat]['姓名'].values[0]
                    st.markdown(f"#### 📈 {analysis_seat} 號 【{stud_name}】 學習軌跡綜合分析")
                    
                    fig_score, ax_score = plt.subplots(figsize=(8, 4.5))
                    sub_cols = ["國文", "英文", "數學", "自然", "社會(含領域融合)"]
                    
                    for sub in sub_cols:
                        if sub in df_history.columns:
                            ax_score.plot(df_history["考試名稱"], df_history[sub], marker='o', linewidth=2, label=f"{sub} PR")
                    
                    ax_score.set_title("各科 PR 值走勢圖 (跨學期真實能力追蹤)", fontsize=14, fontweight='bold')
                    ax_score.set_ylabel("PR 值 (越大表現越優異)", fontsize=12)
                    ax_score.set_ylim(0, 105)
                    ax_score.grid(True, linestyle=':', alpha=0.6)
                    ax_score.legend(loc='lower left', ncol=5, fontsize=9)
                    plt.xticks(rotation=15)
                    st.pyplot(fig_score)
                    
                    fig_rank, ax_rank = plt.subplots(figsize=(8, 3.5))
                    if "名次" in df_history.columns:
                        ax_rank.plot(df_history["考試名稱"], df_history["名次"], marker='s', color='#E67E22', linewidth=2.5, label="個人排名")
                        ax_rank.set_title("名次晉升軌跡 (曲線往上代表進步)", fontsize=14, fontweight='bold')
                        ax_rank.set_ylabel("名次", fontsize=12)
                        ax_rank.grid(True, linestyle=':', alpha=0.6)
                        ax_rank.invert_yaxis() 
                        max_rank = int(df_history["名次"].max()) + 3
                        ax_rank.set_ylim(max_rank, 1)
                        plt.xticks(rotation=15)
                        st.pyplot(fig_rank)
                    
                    with st.expander("👀 檢視完整歷史數據清單"):
                        st.dataframe(df_history)
                else:
                    st.warning("找不到該位同學的歷史紀錄。")

    with tab3:
        st.markdown("### 🎯 國中教育會考 ─ PR 落點對照標準")
        st.info("平時段考分數會因為難易度而起伏，但 **PR 值 (百分等級)** 代表你在全校的相對排名位置，是預測會考落點最精準的數字！")
        
        col_table1, col_table2, col_table3 = st.columns(3)
        with col_table1:
            st.success("**🏆 A 級 (精熟)**\n\n**A++**：PR 91 ～ 99\n**A+**：PR 85 ～ 90\n**A**：PR 75 ～ 84\n\n*(目標拿 A 者，各科請維持 PR 75 以上)*")
        with col_table2:
            st.warning("**💪 B 級 (基礎)**\n\n**B++**：PR 60 ～ 74\n**B+**：PR 45 ～ 59\n**B**：PR 15 ～ 44\n\n*(多數同學落點，拚進步的關鍵)*")
        with col_table3:
            st.error("**⚠️ C 級 (待加強)**\n\n**C**：PR 14 以下\n\n*(基礎觀念需要重新補強)*")

        st.markdown("---")

        if st.session_state.role == "teacher":
            st.markdown("#### 👨‍🏫 導師專屬：班級會考預估落點總表結算")
            target_class = st.text_input("請輸入欲查詢的班級代碼 (輸入 01 可抓取所有 1 班；輸入 701 僅抓取 701 班)：", value="01")
            
            if st.button("🚀 結算該班會考預估落點 (跨學期 PR 平均)", type="primary"):
                with st.spinner(f"正在全校資料庫中撈取符合 {target_class} 的學生，計算會考等級..."):
                    df_pr, excel_pr = generate_pr_report_excel(df_menu, target_class)
                    if df_pr is not None and not df_pr.empty:
                        st.success(f"✅ 成功結算符合 {target_class} 的學生共 {len(df_pr)} 位落點！")
                        st.download_button(
                            label=f"📥 下載【{target_class}班會考預估落點 Excel】",
                            data=excel_pr,
                            file_name=f"班級{target_class}_會考預估落點總表.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.dataframe(df_pr)
                    else:
                        st.error(f"❌ 找不到包含代碼 '{target_class}' 的班級資料。")
        else:
            st.markdown("#### 🎯 你的個人專屬預估落點")
            if st.button("🚀 分析我的會考落點", type="primary"):
                if 'student_seats' in locals() and st.session_state.user_id in student_seats:
                    df_history = fetch_student_history(df_menu, st.session_state.user_id)
                    if not df_history.empty:
                        my_pr = {
                            "國文": df_history["國文"].mean(), "英文": df_history["英文"].mean(),
                            "數學": df_history["數學"].mean(), "自然": df_history["自然"].mean(),
                            "社會": df_history["社會(含領域融合)"].mean() if "社會(含領域融合)" in df_history.columns else np.nan
                        }
                        
                        st.write("根據你歷次段考的 PR 值平均，為你預估的會考落點如下：")
                        cols = st.columns(5)
                        for i, (sub, pr) in enumerate(my_pr.items()):
                            if pd.notna(pr):
                                cols[i].metric(label=f"{sub} (平均 PR: {pr:.1f})", value=pr_to_level(pr))
                        
                        total_pr = pd.Series(list(my_pr.values())).mean()
                        if pd.notna(total_pr):
                            st.info(f"✨ **你的五科綜合預估落點：{pr_to_level(total_pr)}** (總平均 PR: {total_pr:.1f})")
                    else:
                        st.warning("目前還沒有足夠的歷史資料來為你進行分析喔！")
else:
    st.error("❌ 無法載入您的主控總表，請確認網址權限設定。")
